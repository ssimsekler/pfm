"""Full data export to XLSX (Decision #29).

Exports config, master and transactional data. Two modes:
  - single workbook with one worksheet per entity (default; returned as bytes),
  - separate .xlsx files written to a server-side folder (opt-in via `to_folder`).

Uses openpyxl. Column order follows the model's table columns; UUID/date values
are stringified for portability.
"""

import io
import os
import uuid as uuid_lib
from datetime import date, datetime
from decimal import Decimal

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

import app.models as m

# (worksheet name, model) — ordered by group. Sheet names <= 31 chars.
EXPORT_TABLES = [
    # Config / meta
    ("app_config", m.AppConfig),
    ("id_sequence", m.IdSequence),
    ("code_list", m.CodeList),
    ("code_value", m.CodeValue),
    ("llm_provider", m.LlmProvider),
    ("feature_llm_binding", m.FeatureLlmBinding),
    ("integration_endpoint", m.IntegrationEndpoint),
    ("audit_log", m.AuditLog),
    ("event_outbox", m.EventOutbox),
    # Security
    ("household", m.Household),
    ("app_user", m.AppUser),
    ("role", m.Role),
    ("user_role", m.UserRole),
    # Reference
    ("currency", m.Currency),
    ("country", m.Country),
    ("institution", m.Institution),
    ("currency_rate", m.CurrencyRate),
    # Master
    ("account", m.Account),
    ("partner", m.Partner),
    ("beneficiary", m.Beneficiary),
    ("expense_category", m.ExpenseCategory),
    ("cash_flow_item", m.CashFlowItem),
    ("tag", m.Tag),
    ("entity_tag", m.EntityTag),
    ("recurrence_profile", m.RecurrenceProfile),
    ("holiday_calendar", m.HolidayCalendar),
    ("holiday_calendar_day", m.HolidayCalendarDay),
    ("installment_plan", m.InstallmentPlan),
    ("installment_schedule", m.InstallmentSchedule),
    ("loan", m.Loan),
    ("amortization_schedule", m.AmortizationSchedule),
    ("goal", m.Goal),
    ("investment_holding", m.InvestmentHolding),
    ("valuation_history", m.ValuationHistory),
    ("budget", m.Budget),
    ("budget_line", m.BudgetLine),
    ("categorization_rule", m.CategorizationRule),
    ("notification", m.Notification),
    # Transactional
    ("transaction", m.Transaction),
    ("transaction_split", m.TransactionSplit),
    ("transfer_group", m.TransferGroup),
    ("attachment", m.Attachment),
    # Imports
    ("document_import", m.DocumentImport),
    ("document_import_row", m.DocumentImportRow),
]


def _cell(value):
    if value is None:
        return ""
    if isinstance(value, (uuid_lib.UUID,)):
        return str(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (dict, list)):
        import json
        return json.dumps(value, default=str)
    return value


def _rows_for(db: Session, model):
    columns = [c.name for c in model.__table__.columns]
    objects = db.execute(select(model)).scalars().all()
    data = []
    for obj in objects:
        data.append([_cell(getattr(obj, col)) for col in columns])
    return columns, data


def build_workbook(db: Session) -> bytes:
    """Single workbook, one sheet per entity."""
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet
    for sheet_name, model in EXPORT_TABLES:
        columns, data = _rows_for(db, model)
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.append(columns)
        for row in data:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_separate_files(db: Session, folder: str) -> list[str]:
    """Write one .xlsx per entity into `folder` (server-side). Returns paths."""
    os.makedirs(folder, exist_ok=True)
    paths = []
    for sheet_name, model in EXPORT_TABLES:
        columns, data = _rows_for(db, model)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        ws.append(columns)
        for row in data:
            ws.append(row)
        path = os.path.join(folder, f"{sheet_name}.xlsx")
        wb.save(path)
        paths.append(path)
    return paths