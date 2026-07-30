"""Full data import from XLSX (round-trip of export_data, Decision #29).

Reloads a clean instance from a workbook produced by `export_data.build_workbook`.
Semantics: for each entity, existing rows are removed and the sheet content is
written verbatim, preserving original UUID primary keys so foreign-key
references resolve.

Design notes / caveats handled here:
  - Insert order follows EXPORT_TABLES (parents before children); delete order is
    the reverse. Self-referencing and circular FKs (beneficiary/expense_category
    parent_id; transaction <-> transfer_group) are handled by deferring
    constraints for the duration of the transaction (PostgreSQL).
  - Cell values are coerced back to column types (inverse of export_data._cell):
    UUID strings -> UUID, ISO strings -> date/datetime, numeric strings -> Decimal,
    JSON strings -> dict/list, "" -> NULL for nullable columns.
  - The whole load runs in a single transaction; on any error it rolls back.
"""

import io
import json
import uuid as uuid_lib
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import (
    Boolean,
    Date,
    DateTime,
    Integer,
    Numeric,
    SmallInteger,
    text,
)
from sqlalchemy.dialects.postgresql import UUID as PGUUID
from sqlalchemy.orm import Session

from app.services.export_data import EXPORT_TABLES


class ImportError_(Exception):
    """Raised when the workbook cannot be imported."""


def _is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _coerce(value, column):
    """Inverse of export_data._cell: convert a cell value to the column's type."""
    coltype = column.type
    nullable = column.nullable

    if _is_blank(value):
        if nullable:
            return None
        # Non-nullable: fall through so type coercion can produce a default-ish
        # value only where sensible; otherwise return "" / 0 as appropriate.

    # UUID
    if isinstance(coltype, PGUUID):
        if _is_blank(value):
            return None if nullable else value
        return uuid_lib.UUID(str(value))

    # Booleans
    if isinstance(coltype, Boolean):
        if _is_blank(value):
            return None if nullable else False
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in ("1", "true", "t", "yes", "y")

    # Integers
    if isinstance(coltype, (Integer, SmallInteger)):
        if _is_blank(value):
            return None if nullable else 0
        return int(float(value))

    # Numeric / Decimal — reconstruct exactly via str to avoid float artifacts.
    if isinstance(coltype, Numeric):
        if _is_blank(value):
            return None if nullable else Decimal("0")
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError) as exc:
            raise ImportError_(f"Invalid numeric value {value!r} for {column.name}") from exc

    # Dates / datetimes
    if isinstance(coltype, DateTime):
        if _is_blank(value):
            return None
        if isinstance(value, datetime):
            return value
        return datetime.fromisoformat(str(value))
    if isinstance(coltype, Date):
        if _is_blank(value):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return date.fromisoformat(str(value))

    # JSON columns are stringified on export; re-parse if it looks like JSON.
    type_name = coltype.__class__.__name__.upper()
    if "JSON" in type_name:
        if _is_blank(value):
            return None
        if isinstance(value, (dict, list)):
            return value
        try:
            return json.loads(value)
        except (json.JSONDecodeError, TypeError):
            return value

    # Strings / text and everything else
    if _is_blank(value):
        return None if nullable else ""
    return value


def _rows_from_sheet(ws, model):
    """Yield dict(column -> coerced value) for each data row in the worksheet."""
    columns_by_name = {c.name: c for c in model.__table__.columns}
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(h) if h is not None else "" for h in row]
            continue
        if all(_is_blank(v) for v in row):
            continue
        record = {}
        for name, raw in zip(header, row):
            col = columns_by_name.get(name)
            if col is None:
                continue  # unknown column in sheet — ignore
            record[name] = _coerce(raw, col)
        yield record


def import_workbook(db: Session, content: bytes) -> dict:
    """Wipe and reload all entities from a workbook. Runs in one transaction.

    Returns a summary: {table: rows_written}.
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_names = set(wb.sheetnames)
    summary: dict[str, int] = {}

    is_postgres = db.bind.dialect.name == "postgresql"

    try:
        if is_postgres:
            db.execute(text("SET CONSTRAINTS ALL DEFERRED"))

        # Delete children before parents (reverse of insert order).
        for sheet_name, model in reversed(EXPORT_TABLES):
            db.execute(model.__table__.delete())

        # Insert parents before children.
        for sheet_name, model in EXPORT_TABLES:
            written = 0
            if sheet_name[:31] in sheet_names:
                ws = wb[sheet_name[:31]]
                records = list(_rows_from_sheet(ws, model))
                if records:
                    db.execute(model.__table__.insert(), records)
                    written = len(records)
            summary[sheet_name] = written

        db.commit()
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        raise ImportError_(str(exc)) from exc
    finally:
        wb.close()

    return summary